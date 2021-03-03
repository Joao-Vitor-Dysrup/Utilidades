import openpyxl
from openpyxl.styles import PatternFill


class ExcelManager:
    def __init__(self, filepath):
        self.filepath = filepath
        self.workbook = self.get_workbook()
        self.sheet = self.workbook.active
        self.num_rows = self.sheet.max_row
        self.num_columns = self.sheet.max_column
        self.attributes = self.identify_attributes()

    def identify_attributes(self):
        return [self.read_cell(1, j) for j in range(1, self.num_columns + 1)]

    def read_row(self, row_number):
        return {self.attributes[j - 1]: self.read_cell(row_number, j)
                for j in range(1, self.num_columns + 1)}

    def read_cell(self, row_number, column_number):
        return self.sheet.cell(row=row_number, column=column_number).value

    def write_cell(self, row_number, column_number, value):
        if not self.is_cell_empty(row_number, column_number):
            print(f'Cannot write to'
                  f' cell ({row_number}, {column_number}). It is not empty!')
            return
        self.sheet.cell(row=row_number, column=column_number).value = value

    def write_row(self, row_number, list_of_values):
        if not self.is_row_empty(row_number):
            print(f'Cannot write to row {row_number}. It does not have empty cells.')
            return
        for position, value in enumerate(list_of_values):
            self.write_cell(
                row_number=row_number,
                column_number=position + 1,
                value=value)

    def is_row_empty(self, row_number):
        for j in range(1, self.num_columns + 1):
            cell_value = self.read_cell(row_number, j)
            print(cell_value)
            if not self.is_cell_empty(row_number, j):
                return False
        return True

    def is_cell_empty(self, row_number, column_number):
        return self.read_cell(row_number, column_number) != 'None'

    def read_all(self):
        """
        Reads the entire excel file, storing each row as a dictionary
        and putting each dictionary inside a list (of rows)
        :return: list of dictionaries
        """
        return [self.read_row(i) for i in range(2, self.num_rows + 1)]

    def append(self, list_of_values):
        self.sheet.append(list_of_values)

    def save(self):
        self.align_text_center()
        self.workbook.save(self.filepath)

    def get_workbook(self):
        try:
            openpyxl.load_workbook(self.filepath)
        except FileNotFoundError:
            _ = openpyxl.Workbook()
            _.save(self.filepath)
        finally:
            return openpyxl.load_workbook(self.filepath)

    def align_text_center(self):
        for col in self.sheet.columns:
            for cell in col:
                alignment_obj = cell.alignment.copy(horizontal='center', vertical='center')
                cell.alignment = alignment_obj

    def color_row(self, color_hexa, row_number, num_cols):
        for rows in self.sheet.iter_rows(
                min_row=row_number,
                max_row=row_number,
                min_col=1, max_col=num_cols):
            for cell in rows:
                cell.fill = PatternFill(
                    fgColor=color_hexa,
                    bgColor=color_hexa,
                    fill_type="solid",
                    start_color=color_hexa,
                    end_color=color_hexa,
                )

    def write_attributes(self, attributes):
        self.color_row('FF9900', 1, len(attributes))
        self.write_row(row_number=1, list_of_values=attributes)
